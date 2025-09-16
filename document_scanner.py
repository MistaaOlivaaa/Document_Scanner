import cv2
import numpy as np
from PIL import Image, ImageEnhance
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
import argparse
import sys
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
from datetime import datetime

class DocumentScanner:
    def __init__(self):
        self.original_image = None
        self.processed_image = None
        self.document_contour = None
        
    def load_image(self, image_path):
        """Load an image from the specified path"""
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image not found: {image_path}")
        
        self.original_image = cv2.imread(image_path)
        if self.original_image is None:
            raise ValueError("Could not load image. Please check the file format.")
        
        print(f"✅ Image loaded successfully: {image_path}")
        print(f"📐 Image dimensions: {self.original_image.shape[1]}x{self.original_image.shape[0]}")
        return self.original_image
    
    def detect_document_edges(self, image=None):
        """Detect document edges using contour detection"""
        if image is None:
            image = self.original_image
        
        if image is None:
            raise ValueError("No image loaded. Please load an image first.")
        
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Apply Gaussian blur to reduce noise
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        
        # Apply edge detection
        edged = cv2.Canny(blurred, 75, 200)
        
        # Find contours
        contours, _ = cv2.findContours(edged, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]
        
        # Find the document contour (should be rectangular)
        document_contour = None
        for contour in contours:
            # Approximate the contour
            peri = cv2.arcLength(contour, True)
            approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
            
            # If we found a 4-point contour, we likely found the document
            if len(approx) == 4:
                document_contour = approx
                break
        
        if document_contour is None:
            print("⚠️  Could not detect document edges automatically. Using full image.")
            # Use the entire image as fallback
            h, w = image.shape[:2]
            document_contour = np.array([
                [0, 0], [w, 0], [w, h], [0, h]
            ], dtype=np.float32).reshape(4, 1, 2)
        else:
            print("✅ Document edges detected successfully!")
        
        self.document_contour = document_contour
        return document_contour
    
    def order_points(self, pts):
        """Order points in top-left, top-right, bottom-right, bottom-left order"""
        rect = np.zeros((4, 2), dtype="float32")
        
        # Sum and difference to find corners
        s = pts.sum(axis=1)
        diff = np.diff(pts, axis=1)
        
        # Top-left point has smallest sum
        rect[0] = pts[np.argmin(s)]
        # Bottom-right point has largest sum
        rect[2] = pts[np.argmax(s)]
        # Top-right point has smallest difference
        rect[1] = pts[np.argmin(diff)]
        # Bottom-left point has largest difference
        rect[3] = pts[np.argmax(diff)]
        
        return rect
    
    def perspective_transform(self, image=None, contour=None):
        """Apply perspective transformation to get top-down view"""
        if image is None:
            image = self.original_image
        if contour is None:
            contour = self.document_contour
            
        if image is None or contour is None:
            raise ValueError("Image and contour are required for perspective transformation")
        
        # Reshape contour and order points
        pts = contour.reshape(4, 2)
        rect = self.order_points(pts)
        
        # Calculate the width and height of the new image
        (tl, tr, br, bl) = rect
        widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
        widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
        maxWidth = max(int(widthA), int(widthB))
        
        heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
        heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
        maxHeight = max(int(heightA), int(heightB))
        
        # Define destination points for the transform
        dst = np.array([
            [0, 0],
            [maxWidth - 1, 0],
            [maxWidth - 1, maxHeight - 1],
            [0, maxHeight - 1]
        ], dtype="float32")
        
        # Calculate perspective transform matrix and apply it
        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
        
        print("✅ Perspective transformation applied successfully!")
        return warped
    
    def enhance_image(self, image, enhancement_mode='adaptive'):
        """Enhance the image to look like a clean scan"""
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        if enhancement_mode == 'adaptive':
            # Apply adaptive thresholding for better results
            enhanced = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                cv2.THRESH_BINARY, 11, 2
            )
        elif enhancement_mode == 'otsu':
            # Apply Otsu's thresholding
            _, enhanced = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        else:  # simple threshold
            # Apply simple thresholding
            _, enhanced = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY)
        
        # Optional: Apply morphological operations to clean up the image
        kernel = np.ones((2, 2), np.uint8)
        enhanced = cv2.morphologyEx(enhanced, cv2.MORPH_CLOSE, kernel)
        enhanced = cv2.morphologyEx(enhanced, cv2.MORPH_OPEN, kernel)
        
        print(f"✅ Image enhanced using {enhancement_mode} method!")
        return enhanced
    
    def save_image(self, image, output_path, quality=95):
        """Save the processed image"""
        # Convert from OpenCV BGR to RGB for PIL
        if len(image.shape) == 3:
            image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        else:
            image_rgb = image
        
        # Convert to PIL Image
        pil_image = Image.fromarray(image_rgb)
        
        # Save with high quality
        pil_image.save(output_path, quality=quality, optimize=True)
        print(f"✅ Image saved: {output_path}")
        
    def save_as_pdf(self, image, output_path, page_size='A4'):
        """Save the processed image as PDF"""
        # Convert from OpenCV format to PIL
        if len(image.shape) == 3:
            image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        else:
            image_rgb = image
        
        pil_image = Image.fromarray(image_rgb)
        
        # Save as PDF
        if page_size == 'A4':
            pil_image.save(output_path, "PDF", resolution=100.0, save_all=True)
        else:
            pil_image.save(output_path, "PDF", resolution=100.0, save_all=True)
        
        print(f"✅ PDF saved: {output_path}")
    
    def save_as_excel(self, image, output_path, input_filename=""):
        """Save the processed image as Excel file with metadata"""
        try:
            # Convert from OpenCV format to PIL and save as temporary image
            if len(image.shape) == 3:
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            else:
                image_rgb = image
            
            pil_image = Image.fromarray(image_rgb)
            
            # Save temporary image file for Excel embedding
            temp_image_path = "temp_scan.png"
            pil_image.save(temp_image_path, quality=95, optimize=True)
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scanned Document"
            
            # Add metadata header
            ws['A1'] = "Document Scan Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add scan information
            ws['A3'] = "Scan Date:"
            ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A4'] = "Original File:"
            ws['B4'] = input_filename if input_filename else "Unknown"
            ws['A5'] = "Image Dimensions:"
            ws['B5'] = f"{image.shape[1]} x {image.shape[0]} pixels"
            ws['A6'] = "Processing Status:"
            ws['B6'] = "Successfully Processed"
            
            # Style the metadata
            for row in range(3, 7):
                ws[f'A{row}'].font = Font(bold=True)
            
            # Add the scanned image
            img = ExcelImage(temp_image_path)
            
            # Resize image to fit in Excel (max width: 600px)
            max_width = 600
            if img.width > max_width:
                ratio = max_width / img.width
                img.width = max_width
                img.height = int(img.height * ratio)
            
            # Position image starting from row 8
            ws.add_image(img, 'A8')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            
            # Save Excel file
            wb.save(output_path)
            
            # Clean up temporary image
            if os.path.exists(temp_image_path):
                os.remove(temp_image_path)
            
            print(f"✅ Excel file saved: {output_path}")
            
        except Exception as e:
            print(f"❌ Error saving Excel file: {str(e)}")
            # Clean up temporary image in case of error
            if os.path.exists("temp_scan.png"):
                os.remove("temp_scan.png")
    
    def process_document(self, input_path, output_dir="output", 
                        enhancement_mode='adaptive', save_format='all'):
        """Complete document processing pipeline"""
        try:
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Step 1: Load image
            print("📸 Loading image...")
            self.load_image(input_path)
            
            # Step 2: Detect document edges
            print("🔍 Detecting document edges...")
            self.detect_document_edges()
            
            # Step 3: Apply perspective transformation
            print("✂️ Applying perspective transformation...")
            warped = self.perspective_transform()
            
            # Step 4: Enhance image
            print("🌟 Enhancing image quality...")
            enhanced = self.enhance_image(warped, enhancement_mode)
            
            # Step 5: Save results
            base_name = os.path.splitext(os.path.basename(input_path))[0]
            
            if save_format in ['image', 'all']:
                image_path = os.path.join(output_dir, f"{base_name}_scanned.png")
                self.save_image(enhanced, image_path)
            
            if save_format in ['pdf', 'all']:
                pdf_path = os.path.join(output_dir, f"{base_name}_scanned.pdf")
                self.save_as_pdf(enhanced, pdf_path)
            
            if save_format in ['excel', 'all']:
                excel_path = os.path.join(output_dir, f"{base_name}_scanned.xlsx")
                self.save_as_excel(enhanced, excel_path, os.path.basename(input_path))
            
            self.processed_image = enhanced
            print("🎉 Document scanning completed successfully!")
            
            return enhanced
            
        except Exception as e:
            print(f"❌ Error processing document: {str(e)}")
            return None
    
    def preview_detection(self, input_path, output_path="preview.png"):
        """Create a preview showing detected edges"""
        self.load_image(input_path)
        contour = self.detect_document_edges()
        
        # Draw the contour on the original image
        preview = self.original_image.copy()
        cv2.drawContours(preview, [contour], -1, (0, 255, 0), 3)
        
        # Add corner points
        for point in contour.reshape(4, 2):
            cv2.circle(preview, tuple(point.astype(int)), 10, (255, 0, 0), -1)
        
        self.save_image(preview, output_path)
        print(f"📋 Preview saved: {output_path}")
        return preview

def main():
    parser = argparse.ArgumentParser(description="Mobile-friendly Document Scanner")
    parser.add_argument("input", help="Input image path")
    parser.add_argument("-o", "--output", default="output", help="Output directory")
    parser.add_argument("-e", "--enhancement", choices=['adaptive', 'otsu', 'simple'], 
                       default='adaptive', help="Enhancement method")
    parser.add_argument("-f", "--format", choices=['image', 'pdf', 'excel', 'all'], 
                       default='all', help="Output format")
    parser.add_argument("-p", "--preview", action="store_true", 
                       help="Generate preview showing detected edges")
    
    args = parser.parse_args()
    
    scanner = DocumentScanner()
    
    if args.preview:
        print("📋 Generating preview...")
        scanner.preview_detection(args.input, "detection_preview.png")
    
    print("🚀 Starting document scanning process...")
    result = scanner.process_document(
        args.input, 
        args.output, 
        args.enhancement, 
        args.format
    )
    
    if result is not None:
        print("✨ Scanning completed! Check the output directory for results.")
    else:
        print("💥 Scanning failed. Please check the input image and try again.")

if __name__ == "__main__":
    # Example usage if run without arguments
    if len(sys.argv) == 1:
        print("📱 Document Scanner - Usage Examples:")
        print("python document_scanner.py image.jpg")
        print("python document_scanner.py image.jpg -o scanned_docs -e adaptive -f all")
        print("python document_scanner.py image.jpg -f excel")
        print("python document_scanner.py image.jpg --preview")
        print("\nFor help: python document_scanner.py -h")
    else:
        main()
